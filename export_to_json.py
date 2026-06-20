"""Export Performance_Arena_Dataset.xlsx back to data.js.

Workflow:
    1. Edit Performance_Arena_Dataset.xlsx in Excel.
    2. Save the workbook.
    3. Run: python export_to_json.py
    4. Run: python validate_data.py
    5. Run: node test_prototype.js
    6. In the browser, reset state:
       localStorage.removeItem('arena_state_v6'); location.reload();
"""
from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Performance_Arena_Dataset.xlsx"
DATA_JS = ROOT / "data.js"

ENTITY_SHEETS = [
    "Users", "Teams", "Processes", "KPI_Master",
    "Performance_Data", "Daily_Agent_Score", "Agent_Current",
    "Leaderboard", "Points_Ledger", "XP_Ledger",
    "Missions", "Mission_Assignments",
    "Challenges", "Challenge_Participants", "Challenge_Results",
    "Badges", "Agent_Badges",
    "Rewards", "Reward_Redemptions",
    "Communications", "Communication_Status",
    "Learning_Modules", "Learning_Assignments", "Learning_Completion_Status",
    "PKT_Assessments", "PKT_Questions", "PKT_Attempts",
    "SLA_Commercial_Rules", "Penalty_Reward_Slabs",
    "Commercial_Exposure", "Commercial_Verification",
    "What_If_Scenarios",
    "Coaching", "Recognition",
    "Learning_Points_Rules", "TL_Manager_Verification",
]


def jsonable(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("{") or text.startswith("["):
            try:
                return json.loads(text)
            except Exception:
                return value
    return value


def sheet_records(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if headers == ["_empty"]:
        return []
    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        rec: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            rec[header] = jsonable(value)
        out.append(rec)
    return out


def main():
    if not XLSX.exists():
        raise SystemExit(f"Workbook not found: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    bundle = {}
    missing = []
    total = 0
    for sheet in ENTITY_SHEETS:
        if sheet not in wb.sheetnames:
            missing.append(sheet)
            bundle[sheet] = []
            continue
        rows = sheet_records(wb[sheet])
        bundle[sheet] = rows
        total += len(rows)
    DATA_JS.write_text("window.SEED_DATA = " + json.dumps(bundle, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(f"Wrote {DATA_JS.name}: {len(bundle)} entities, {total} rows")
    if missing:
        print("WARNING missing sheets: " + ", ".join(missing))


if __name__ == "__main__":
    main()
