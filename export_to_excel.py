"""Export Performance Arena seed data from data.js to an editable Excel workbook.

Creates:
    Performance_Arena_Dataset.xlsx

The workbook is intentionally Excel-first: business users can edit data in
sheets, save, then run export_to_json.py to regenerate data.js.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
DATA_JS = ROOT / "data.js"
OUT_XLSX = ROOT / "Performance_Arena_Dataset.xlsx"

ENTITY_SHEETS = [
    "Users", "Teams", "Processes", "KPI_Master",
    "Performance_Data", "Daily_Agent_Score", "Agent_Current",
    "Leaderboard", "Points_Ledger", "XP_Ledger",
    "Missions", "Mission_Assignments",
    "Challenges", "Challenge_Participants", "Challenge_Results",
    "Badges", "Agent_Badges",
    "Rewards", "Reward_Redemptions",
    "Communications", "Communication_Status",
    "Learning_Modules", "Learning_Assignments", "Learning_Completion_Status",
    "PKT_Assessments", "PKT_Questions", "PKT_Attempts",
    "SLA_Commercial_Rules", "Penalty_Reward_Slabs",
    "Commercial_Exposure", "Commercial_Verification",
    "What_If_Scenarios",
    "Coaching", "Recognition",
    "Learning_Points_Rules", "TL_Manager_Verification",
]

SHEET_DOCS: Dict[str, Dict[str, str]] = {
    "Users": {"purpose": "All demo users: 1 Manager, 5 Team Leads, and agents.", "editable": "Name, Location, Level, ArenaPoints, Status, Avatar", "generated": "UserID relationships and manager/team/process links should remain stable.", "ids": "UserID, TeamID, ProcessID, ManagerID"},
    "Teams": {"purpose": "Team/squad master used for TL span and Manager rollups.", "editable": "TeamName, Shift, Location", "generated": "TeamID and leadership links are referential keys.", "ids": "TeamID, ProcessID, TeamLeadID, ManagerID"},
    "Processes": {"purpose": "Process lines for the call-centre prototype.", "editable": "ProcessName, ProcessType, Description", "generated": "ProcessID is referenced by Users, Teams, and KPI applicability.", "ids": "ProcessID"},
    "KPI_Master": {"purpose": "Approved call-centre KPI catalogue and thresholds.", "editable": "KPI_Name, KPI_Type, Unit, Direction, Target, thresholds, Weightage, Description", "generated": "KPI_ID should remain stable because it drives all performance/commercial rows.", "ids": "KPI_ID"},
    "Performance_Data": {"purpose": "Daily KPI rows per user; drives Agent scorecards and historical trends.", "editable": "Actual, Target, Score, Status, Points_Earned, Volume", "generated": "Date/UserID/TeamID/ProcessID/KPI_ID should remain referentially valid.", "ids": "Date, UserID, TeamID, ProcessID, KPI_ID"},
    "Daily_Agent_Score": {"purpose": "Daily weighted score and ranks per agent.", "editable": "PerformanceScore, RAGStatus, Points_Earned, ranks", "generated": "Normally recalculated from Performance_Data.", "ids": "Date, UserID, TeamID, ProcessID"},
    "Agent_Current": {"purpose": "Current snapshot used on Agent Home.", "editable": "PerformanceScore, RAGStatus, balances, ranks", "generated": "Normally derived from latest Daily_Agent_Score and ledgers.", "ids": "UserID, TeamID, ProcessID"},
    "Leaderboard": {"purpose": "Snapshot ranking rows by team/process/account.", "editable": "Rank, Score, Points, XP, Period", "generated": "Can be regenerated from scores and ledgers.", "ids": "Leaderboard_ID, Scope_ID, UserID"},
    "Points_Ledger": {"purpose": "Spendable Arena Points earn/spend ledger.", "editable": "Points_Delta, Description, Source_Type, Timestamp", "generated": "Balance_After is normally calculated.", "ids": "Ledger_ID, UserID, Source_ID"},
    "XP_Ledger": {"purpose": "Internal level-progress ledger retained for compatibility; UI should say Level Progress.", "editable": "XP_Delta, Description, Source_Type, Timestamp", "generated": "Used for level/league progress.", "ids": "Ledger_ID, UserID, Source_ID"},
    "Missions": {"purpose": "Mission/action catalogue for agents, TLs, and managers.", "editable": "Mission_Name, Type, Description, target, rewards, dates, Status", "generated": "Mission_ID and linked IDs should remain stable.", "ids": "Mission_ID, KPI_ID, Audience_ID, Created_By"},
    "Mission_Assignments": {"purpose": "Per-user mission progress and completion state.", "editable": "Progress, Status, earned rewards, dates", "generated": "Assignment_ID is generated; linked Mission_ID/UserID required.", "ids": "Assignment_ID, Mission_ID, UserID, TeamID"},
    "Challenges": {"purpose": "Challenge catalogue: peer, team, TL issued, and manager issued.", "editable": "Challenge_Name, Type, KPI_ID, dates, Entry_Points, Reward_Pool, Status", "generated": "Challenge_ID and Created_By should remain valid.", "ids": "Challenge_ID, KPI_ID, Created_By"},
    "Challenge_Participants": {"purpose": "Participants and sides for each challenge.", "editable": "Side, Status, Entry_Paid, Joined_Date", "generated": "Participant_ID generated; UserID and Challenge_ID must exist.", "ids": "Participant_ID, Challenge_ID, UserID"},
    "Challenge_Results": {"purpose": "Settled challenge winners and validation outcomes.", "editable": "Winner_UserID, Result_Status, Validated_By, Awarded_Points", "generated": "Should be produced by the challenge settlement/validation flow.", "ids": "Result_ID, Challenge_ID, Winner_UserID"},
    "Badges": {"purpose": "Badge catalogue.", "editable": "Badge_Name, Criteria, bonuses, Icon, Status", "generated": "Badge_ID stable for linked awards.", "ids": "Badge_ID"},
    "Agent_Badges": {"purpose": "Badges earned by users.", "editable": "Earned_Date, Source_Type, Status", "generated": "Award_ID generated from badge earning events.", "ids": "Award_ID, UserID, Badge_ID"},
    "Rewards": {"purpose": "Arena Store reward catalogue.", "editable": "Reward_Name, Category, Points_Required, Approval_Required, Stock, Status", "generated": "Reward_ID stable for redemptions.", "ids": "Reward_ID"},
    "Reward_Redemptions": {"purpose": "Reward redemption and approval state.", "editable": "Status, Approved_By, Approval_Date, Comments", "generated": "Redemption_ID generated from store actions.", "ids": "Redemption_ID, UserID, Reward_ID"},
    "Communications": {"purpose": "Broadcasts/announcements.", "editable": "Title, Message, Audience, Priority, Status", "generated": "Communication_ID stable for acknowledgement status.", "ids": "Communication_ID, Created_By, Audience_ID"},
    "Communication_Status": {"purpose": "Per-user acknowledgement status for broadcasts.", "editable": "Status, Acknowledged_At", "generated": "Status_ID generated.", "ids": "Status_ID, Communication_ID, UserID"},
    "Learning_Modules": {"purpose": "Training/broadcast/PKT module catalogue.", "editable": "Title, Description, Module_Type, Audience, rewards, Status", "generated": "Module_ID stable for assignments.", "ids": "Module_ID, Created_By, Audience_ID"},
    "Learning_Assignments": {"purpose": "Per-user learning assignments.", "editable": "Status, Due_Date, Assigned_Date", "generated": "Assignment_ID generated.", "ids": "Assignment_ID, Module_ID, UserID"},
    "Learning_Completion_Status": {"purpose": "Completion progress for learning modules.", "editable": "Status, Progress, Completed_At, Score", "generated": "Completion_ID generated from training actions.", "ids": "Completion_ID, Assignment_ID, UserID, Module_ID"},
    "PKT_Assessments": {"purpose": "Post-knowledge-test assessment catalogue.", "editable": "Title, Pass_Score, Reward_Points, Status", "generated": "Assessment_ID stable.", "ids": "Assessment_ID, Module_ID"},
    "PKT_Questions": {"purpose": "PKT question bank.", "editable": "Question_Text, Options, Correct_Answer, Explanation, Status", "generated": "Question_ID stable.", "ids": "Question_ID, Assessment_ID"},
    "PKT_Attempts": {"purpose": "PKT attempts by users.", "editable": "Score, Passed, Attempt_Date, Status", "generated": "Attempt_ID generated.", "ids": "Attempt_ID, Assessment_ID, UserID"},
    "SLA_Commercial_Rules": {"purpose": "Commercial rules by KPI.", "editable": "Target, thresholds, penalty/reward parameters, Status", "generated": "Rule_ID and KPI_ID links should remain valid.", "ids": "Rule_ID, KPI_ID"},
    "Penalty_Reward_Slabs": {"purpose": "Penalty/reward slab ranges used by What-If and Commercial views.", "editable": "Slab boundaries and penalty/reward amounts", "generated": "Slab_ID generated; Rule_ID should exist.", "ids": "Slab_ID, Rule_ID"},
    "Commercial_Exposure": {"purpose": "Account and team commercial exposure, revenue, and penalty/reward rollups.", "editable": "Forecast values, penalty/reward, revenue/rate assumptions", "generated": "Entity scope should stay Account or Team; TL team values must stay below Manager account rollup.", "ids": "Snapshot_Date, Entity_Level, Entity_ID, KPI_ID"},
    "Commercial_Verification": {"purpose": "Commercial exposure rows with owner/verifier context.", "editable": "Verification_Status, Verified_By, Comments", "generated": "Verifier_Role, Owner_ID and Entity_ID must preserve scope integrity.", "ids": "Verification_ID, Owner_ID, Entity_ID, KPI_ID"},
    "What_If_Scenarios": {"purpose": "Manager simulator scenarios by KPI and improvement assumption.", "editable": "Improvement_Assumption, projected penalty/reward, Recommended_Team", "generated": "Scenario_ID and current baseline values normally generated.", "ids": "Scenario_ID, KPI_ID"},
    "Coaching": {"purpose": "Coaching queue and notes.", "editable": "Coaching_Note, Status, Due_Date, Resolution", "generated": "Coaching_ID generated.", "ids": "Coaching_ID, UserID, Coach_ID"},
    "Recognition": {"purpose": "Recognition events.", "editable": "Recognition_Type, Message, Points_Awarded, Status", "generated": "Recognition_ID generated.", "ids": "Recognition_ID, UserID, Recognized_By"},
    "Learning_Points_Rules": {"purpose": "Rules for points/level progress from learning actions.", "editable": "Points, XP/Level progress values, Status", "generated": "Rule_ID stable.", "ids": "Rule_ID"},
    "TL_Manager_Verification": {"purpose": "TL/Manager validation workflow items for challenges, rewards, and commercial actions.", "editable": "Verification_Status, Comments, Verified_By", "generated": "Verification_ID generated; owner and target IDs must exist.", "ids": "Verification_ID, Owner_ID, Target_ID"},
}

VALIDATION_LISTS = {
    "Role": ["Agent", "Team Lead", "Manager"],
    "RAG": ["Green", "Amber", "Red"],
    "Risk": ["Green", "Watch", "High", "Critical", "Amber", "Red"],
    "Verification": ["Action Pending", "Pending", "Approved", "Rejected", "Validated", "In Review", "Closed"],
    "Direction": ["Higher", "Lower"],
    "Audience": ["Agent", "Team", "Process", "Account", "All"],
    "ChallengeStatus": ["Draft", "Pending", "Active", "Accepted", "Declined", "Pending Validation", "Settled", "Completed"],
    "RewardStatus": ["Active", "Inactive", "Pending", "Approved", "Rejected", "Redeemed", "Completed"],
    "GenericStatus": ["Active", "Inactive", "Pending", "In Progress", "Completed", "Approved", "Rejected", "Overdue", "Not Started", "Acknowledged", "Settled", "Declined"],
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
README_FILL = PatternFill("solid", fgColor="0F766E")
SECTION_FILL = PatternFill("solid", fgColor="E0F2FE")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_seed_data(path: Path = DATA_JS) -> Dict[str, List[Dict[str, Any]]]:
    text = path.read_text(encoding="utf-8")
    match = re.search(r"window\.SEED_DATA\s*=\s*(.*?);\s*$", text, re.S)
    if not match:
        raise ValueError(f"Could not parse window.SEED_DATA from {path}")
    return json.loads(match.group(1))


def ordered_headers(records: Iterable[Dict[str, Any]]) -> List[str]:
    headers: List[str] = []
    seen = set()
    for rec in records:
        for key in rec.keys():
            if key not in seen:
                headers.append(key)
                seen.add(key)
    return headers


def safe_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def should_apply_validation(sheet_name: str, header: str):
    h = header.lower()
    if header == "Role":
        return "Role"
    if "rag" in h or header in {"RAGStatus"}:
        return "RAG"
    if "risk" in h:
        return "Risk"
    if "verification" in h:
        return "Verification"
    if header == "Direction" or "direction" in h:
        return "Direction"
    if header in {"Audience_Type", "Audience"}:
        return "Audience"
    if sheet_name == "Challenges" and header == "Status":
        return "ChallengeStatus"
    if sheet_name in {"Rewards", "Reward_Redemptions"} and header == "Status":
        return "RewardStatus"
    if header == "Status":
        return "GenericStatus"
    return None


def add_validation(ws, header: str, col_idx: int, sheet_name: str, max_row: int):
    list_key = should_apply_validation(sheet_name, header)
    if not list_key:
        return
    values = VALIDATION_LISTS[list_key]
    formula = '"' + ','.join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Choose a value from the list or leave blank."
    dv.errorTitle = "Invalid value"
    dv.prompt = "Choose from the configured prototype values."
    dv.promptTitle = header
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}2:{col_letter}{max(max_row + 100, 1000)}")


def format_sheet(ws, headers: List[str], rows_count: int):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(1, len(headers)))}{max(1, rows_count + 1)}"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2, max_row=min(rows_count + 1, 250), max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 32)
        # Make common text columns wider but capped.
        if any(token in header.lower() for token in ["description", "message", "comment", "note", "title", "name", "criteria"]):
            width = 28
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 24


def add_readme(wb: Workbook):
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Performance Arena Dataset Workbook"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = README_FILL
    ws.merge_cells("A1:E1")
    intro = [
        ["Purpose", "Editable Excel source of truth for the Performance Arena static prototype."],
        ["Workflow", "1) Edit sheets in Excel. 2) Save workbook. 3) Run: python export_to_json.py. 4) Run: python validate_data.py. 5) Run: node test_prototype.js. 6) Reset browser localStorage."],
        ["Reset browser state", "localStorage.removeItem('arena_state_v6'); location.reload();"],
        ["Important guardrail", "Do not change IDs unless you also update all linked rows. Keep TL commercial rows team-scoped and Manager rows account-scoped."],
    ]
    row = 3
    for k, v in intro:
        ws.cell(row, 1).value = k
        ws.cell(row, 2).value = v
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
    row += 1
    headers = ["Sheet", "Purpose", "Editable fields", "Calculated/generated fields", "Key IDs / notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c)
        cell.value = h
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    row += 1
    for sheet in ENTITY_SHEETS:
        doc = SHEET_DOCS.get(sheet, {})
        vals = [
            sheet,
            doc.get("purpose", "Prototype entity table."),
            doc.get("editable", "Business values and demo labels where applicable."),
            doc.get("generated", "IDs, relationships, and derived values should be changed carefully."),
            doc.get("ids", "See headers in sheet."),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row, c)
            cell.value = val
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    widths = [28, 46, 44, 48, 42]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A9"


def main():
    data = load_seed_data()
    wb = Workbook()
    add_readme(wb)
    for sheet_name in ENTITY_SHEETS:
        records = data.get(sheet_name, []) or []
        ws = wb.create_sheet(sheet_name[:31])
        headers = ordered_headers(records)
        if not headers:
            headers = ["_empty"]
        ws.append(headers)
        for rec in records:
            ws.append([safe_value(rec.get(h)) for h in headers])
        for col_idx, header in enumerate(headers, 1):
            add_validation(ws, header, col_idx, sheet_name, len(records) + 1)
        format_sheet(ws, headers, len(records))
        ws["A1"].comment = Comment(
            f"{SHEET_DOCS.get(sheet_name, {}).get('purpose', 'Prototype entity table.')}\n\n"
            f"Run python export_to_json.py after editing this workbook.",
            "Performance Arena",
        )
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX.name} with {len(ENTITY_SHEETS)} data sheets + README")


if __name__ == "__main__":
    main()
